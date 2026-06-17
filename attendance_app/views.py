# attendance_app/views.py
import qrcode
import io
import hmac
import hashlib
import time
import base64
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from .models import Session, Attendance, Subject, Group, Profile
from .forms import SessionForm, LoginForm, ReportForm, UserUpdateForm, StudentGroupForm
from django.urls import reverse

from .forms import SessionForm, LoginForm, ReportForm

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.db.models import Count, Q
from datetime import datetime

import random
from django.utils import timezone
from datetime import timedelta

# --- Криптографические функции ---

def generate_signed_payload(session_id: str, secret_key: bytes = None) -> str:
    
    # Генерирует строку для QR-кода вида: session_id|timestamp|hmac_signature
    
    if secret_key is None:
        # Определяем секретный ключ. Eсли не передан явно, используется стандартный SECRET_KEY Django
        secret_key = settings.SECRET_KEY.encode('utf-8')
    
    timestamp = int(time.time()) # Получаем текущую временную метку в секундах
    message = f"{session_id}|{timestamp}".encode('utf-8') # Формируем сообщение для подписи
    signature = hmac.new(secret_key, message, hashlib.sha256).hexdigest() # Вычисляем HMAC-SHA256 подпись от сообщения с использованием секретного ключа
    return f"{session_id}|{timestamp}|{signature}" # Возвращаем итоговый payload

def verify_signed_payload(payload: str, max_age_seconds: int = 30, secret_key: bytes = None) -> bool:
    
    #Проверяет подпись и срок действия payload.
    #Возвращает True, если данные корректны и не устарели.
    
    if secret_key is None:
        secret_key = settings.SECRET_KEY.encode('utf-8') #  Определяем секретный ключ
    
    try:
        parts = payload.split('|')  # Разбираем строку payload на составные части
        if len(parts) != 3:
            return False
        session_id, timestamp_str, signature = parts
        timestamp = int(timestamp_str) # приводим временную метку к целому числу
    except (ValueError, TypeError):
        return False

    # Проверка времени
    current_time = int(time.time())
    if current_time - timestamp > max_age_seconds:
        return False

    # Проверка подписи
    message = f"{session_id}|{timestamp}".encode('utf-8')
    expected_signature = hmac.new(secret_key, message, hashlib.sha256).hexdigest()
    return hmac.compare_digest(signature, expected_signature)  # Для сравнения используем compare_digest, устойчивую к атакам по времени.

# --- Представления для преподавателя ---

@login_required
def create_session(request):
    # Доступ только для преподавателей
    if request.user.profile.role != 'teacher':
        return redirect('attendance_app:index')

    if request.method == 'POST':
        form = SessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.teacher = request.user
            session.save()
            return redirect('attendance_app:show_qr', session_id=session.id)
    else:
        form = SessionForm()
    return render(request, 'attendance_app/create_session.html', {'form': form})


def generate_unique_otp():
    """Генерирует 6-значный код, уникальный среди активных сессий."""
    while True:
        code = f"{random.randint(0, 999999):06d}"
        # Проверяем, нет ли уже активной сессии с таким кодом
        if not Session.objects.filter(otp_code=code, otp_expiry__gt=timezone.now()).exists():
            return code


@login_required
def session_qr(request, session_id):
    session = get_object_or_404(Session, pk=session_id)
    if session.teacher != request.user and request.user.profile.role != 'admin':
        return redirect('attendance_app:index')

    now = timezone.now()
    # Если код отсутствует или просрочен, генерируем новый
    if not session.otp_code or not session.otp_expiry or session.otp_expiry < now:
        session.otp_code = generate_unique_otp()
        session.otp_expiry = now + timedelta(seconds=30)
        session.save()

    # Генерация QR с полной ссылкой (для сканирования)
    payload = generate_signed_payload(str(session.id))
    mark_url = request.build_absolute_uri(
        reverse('attendance_app:mark_attendance', kwargs={'session_id': session.id})
    )
    full_url = f"{mark_url}?payload={payload}"
    qr = qrcode.make(full_url)
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    buf.seek(0)
    data_uri = 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode('ascii')

    context = {
        'data_uri': data_uri,
        'session': session,
        'update_interval': 30,
        'otp_code': session.otp_code,
    }
    return render(request, 'attendance_app/show_qr.html', context)




# --- Заглушки для других представлений (пока не трогаем) ---

def index(request):
    user = request.user
    context = {}
    if user.is_authenticated:
        role = user.profile.role
        if role == 'teacher':
            # Сессии преподавателя
            sessions = Session.objects.filter(teacher=user).order_by('-start_time')
            context['sessions'] = sessions
            context['role'] = 'teacher'
        elif role == 'student':
            # Отметки студента
            attendances = Attendance.objects.filter(student=user).select_related('session__subject').order_by('-timestamp')
            context['attendances'] = attendances
            context['role'] = 'student'
        elif role == 'admin':
            context['role'] = 'admin'
    return render(request, 'attendance_app/index.html', context)

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Перенаправляем на главную после успешного входа
                return redirect('attendance_app:index')
            else:
                # Неверный логин или пароль
                return render(request, 'attendance_app/login.html', {
                    'form': form,
                    'error': 'Неверное имя пользователя или пароль.'
                })
    else:
        form = LoginForm()
    return render(request, 'attendance_app/login.html', {'form': form})


def user_logout(request):
    logout(request)
    return redirect('attendance_app:index')


# attendance_app/views.py (добавить в конец)
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from .models import Attendance



@login_required


def mark_attendance(request, session_id=None):
    if request.user.profile.role != 'student':
        return JsonResponse({'error': 'Только студенты могут отмечаться'}, status=403)

    # Параметры
    payload = request.GET.get('payload') or request.POST.get('payload')
    otp = request.GET.get('otp') or request.POST.get('otp')
    session_obj = None

    # Вариант 1: передан payload (сканирование QR)
    if payload:
        if not verify_signed_payload(payload, max_age_seconds=60):
            return render(request, 'attendance_app/mark_attendance.html', {
                'error': 'Недействительный или устаревший QR-код.',
            })
        try:
            payload_session_id = payload.split('|')[0]
            session_obj = Session.objects.get(pk=payload_session_id)
        except (IndexError, Session.DoesNotExist):
            return render(request, 'attendance_app/mark_attendance.html', {
                'error': 'Занятие не найдено.',
            })

    # Вариант 2: передан OTP (ручной ввод)
    elif otp:
        now = timezone.now()
        try:
            session_obj = Session.objects.get(otp_code=otp, otp_expiry__gt=now)
        except Session.DoesNotExist:
            return render(request, 'attendance_app/mark_attendance.html', {
                'error': 'Неверный или устаревший код.',
            })

    else:
        # Показываем форму
        return render(request, 'attendance_app/mark_attendance.html')

    # Проверка группы студента
    if request.user.profile.group != session_obj.subject.group:
        return render(request, 'attendance_app/mark_attendance.html', {
            'error': 'Вы не приписаны к группе этого занятия.',
        })

    # Создание отметки
    attendance, created = Attendance.objects.get_or_create(
        session=session_obj,
        student=request.user,
    )
    if created:
        message = f'Вы успешно отмечены на занятии "{session_obj.subject.title}"'
    else:
        message = f'Вы уже были отмечены ({attendance.timestamp.strftime("%H:%M:%S")})'

    # Очищаем OTP после успешного использования
    if otp:
        session_obj.otp_code = None
        session_obj.otp_expiry = None
        session_obj.save()

    return render(request, 'attendance_app/attendance_result.html', {
        'success': created,
        'message': message,
        'session': session_obj,
    })




@login_required
def manual_mark(request):
    """Страница для ручного ввода кода (без камеры)."""
    return render(request, 'attendance_app/mark_attendance.html', {'session_id': None})

def generate_excel_report(sessions, attendance_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    # Заголовки
    headers = ["Студент", "Группа"]
    for session in sessions:
        headers.append(session.start_time.strftime("%d.%m.%Y %H:%M"))
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Данные
    for student, group_name, marks in attendance_data:
        row = [student, group_name]
        for present in marks:
            row.append("+" if present else "-")
        ws.append(row)

    # Сохраняем в BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output



@login_required
def attendance_report(request):
    if request.user.profile.role not in ['teacher', 'admin']:
        return redirect('attendance_app:index')

    form = ReportForm(request.GET or None)
    if not form.is_valid():
        return render(request, 'attendance_app/report.html', {'form': form})

    group = form.cleaned_data.get('group')
    subject = form.cleaned_data.get('subject')
    date_from = form.cleaned_data.get('date_from')
    date_to = form.cleaned_data.get('date_to')

    sessions = Session.objects.all()
    if subject:
        sessions = sessions.filter(subject=subject)
    if group:
        sessions = sessions.filter(subject__group=group)
    if date_from:
        sessions = sessions.filter(start_time__date__gte=date_from)
    if date_to:
        sessions = sessions.filter(start_time__date__lte=date_to)
    sessions = sessions.order_by('start_time')

    students = User.objects.filter(profile__role='student')
    if group:
        students = students.filter(profile__group=group)

    attendance_data = []
    for student in students:
        student_name = f"{student.last_name} {student.first_name}" if student.last_name else student.username
        group_name = student.profile.group.name if student.profile.group else "—"
        marks = []
        for session in sessions:
            attended = Attendance.objects.filter(session=session, student=student).exists()
            marks.append(attended)
        attendance_data.append((student_name, group_name, marks))

    output = generate_excel_report(sessions, attendance_data)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    return response



from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages

@login_required
def profile_view(request):
    user = request.user
    is_student = user.profile.role == 'student'

    if request.method == 'POST':
        if is_student:
            # Студент может менять только email
            user.email = request.POST.get('email', user.email)
            user.save()
            messages.success(request, 'Email обновлён.')
            return redirect('attendance_app:profile')
        else:
            # Для преподавателей/админов - полная форма
            user_form = UserUpdateForm(request.POST, instance=user)
            if user_form.is_valid():
                user_form.save()
                messages.success(request, 'Профиль обновлён.')
                return redirect('attendance_app:profile')
    else:
        if is_student:
            user_form = None
        else:
            user_form = UserUpdateForm(instance=user)

    context = {
        'user_form': user_form,
        'is_student': is_student,
        'profile': user.profile,
    }
    return render(request, 'attendance_app/profile.html', context)



@login_required
def password_change(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # чтобы сессия не сбросилась
            messages.success(request, 'Пароль изменён.')
            return redirect('attendance_app:profile')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'attendance_app/password_change.html', {'form': form})

@login_required
def student_report(request):
    if request.user.profile.role != 'student':
        return redirect('attendance_app:index')

    student = request.user
    group = student.profile.group
    if not group:
        return render(request, 'attendance_app/student_report.html', {
            'error': 'Вы не приписаны к группе. Обратитесь к администратору.'
        })

    # Даты из GET-параметров
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Предметы группы
    subjects = Subject.objects.filter(group=group)

    # Сессии по этим предметам с фильтром по датам
    sessions = Session.objects.filter(subject__in=subjects).order_by('start_time')
    if date_from:
        sessions = sessions.filter(start_time__date__gte=date_from)
    if date_to:
        sessions = sessions.filter(start_time__date__lte=date_to)

    dates = sorted(list({s.start_time.date() for s in sessions}))

    # Матрица предмет × дата
    rows = []
    for subject in subjects:
        row = [subject.title]
        for date in dates:
            attended = Attendance.objects.filter(
                session__subject=subject,
                session__start_time__date=date,
                student=student
            ).exists()
            row.append("+" if attended else "-")
        rows.append(row)

    student_name = student.get_full_name() or student.username

    # Если нажата кнопка "Скачать Excel"
    if request.GET.get('download') == 'excel':
        output = generate_student_excel_simple(student_name, group.name, dates, rows)
        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'
        return response

    return render(request, 'attendance_app/student_report.html', {
        'group': group,
        'dates': dates,
        'rows': rows,
        'date_from': date_from,
        'date_to': date_to,
    })

def generate_student_excel_simple(student_name, group_name, dates, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    # Заголовок
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dates)+1)
    title_cell = ws.cell(row=1, column=1, value=f"{student_name}, группа {group_name}")
    title_cell.font = Font(bold=True, size=14)

    # Шапка таблицы
    headers = ["Дисциплина"] + [d.strftime("%d.%m.%Y") for d in dates]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for row in rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output